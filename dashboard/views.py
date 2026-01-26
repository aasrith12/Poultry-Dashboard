from __future__ import annotations

import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, time, timezone as dt_timezone, timedelta
from io import BytesIO
from typing import Any
from urllib import request as urlrequest

from django.http import JsonResponse, HttpResponseBadRequest
from django.shortcuts import redirect, render
from django.conf import settings
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods

from .models import Note, Profile, UploadDataset, ChatSession, ChatMessage, ChatAttachment

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None
from .services import bluconsole
from .utils.blu_xml import parse_devices, parse_measurements


def _owner_key(request) -> str:
    return request.session.get("owner_key") or "guest"


def _set_owner_key(request, key: str) -> None:
    request.session["owner_key"] = key


def _blu_creds(request) -> dict | None:
    return request.session.get("blu_creds")


def _require_blu(view_func):
    def wrapper(request, *args, **kwargs):
        if not _blu_creds(request):
            return redirect(f"/login/?next={request.path}")
        return view_func(request, *args, **kwargs)

    return wrapper


def _json_body(request) -> dict[str, Any]:
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return {}


@ensure_csrf_cookie
def home(request):
    profile = Profile.objects.filter(owner_key=_owner_key(request)).first()
    return render(request, "dashboard/home.html", {"profile": profile})


@ensure_csrf_cookie
def faq(request):
    return render(request, "dashboard/faq.html")


@ensure_csrf_cookie
def login_view(request):
    return render(request, "dashboard/login.html")


@ensure_csrf_cookie
def signup(request):
    return render(request, "dashboard/signup.html")


@ensure_csrf_cookie
def profile(request):
    return render(request, "dashboard/profile.html")


@ensure_csrf_cookie
@_require_blu
def sensor_feed(request):
    return render(request, "dashboard/sensor_feed.html")


@ensure_csrf_cookie
@_require_blu
def visualizations(request):
    return render(request, "dashboard/visualizations.html")


@ensure_csrf_cookie
@_require_blu
def ai(request):
    return render(request, "dashboard/ai.html")


@ensure_csrf_cookie
@_require_blu
def ai_chat(request):
    return render(request, "dashboard/ai_chat.html")


@ensure_csrf_cookie
def contact(request):
    return render(request, "dashboard/contact.html")


@require_http_methods(["POST"])
def api_blu_login(request):
    data = _json_body(request)
    uname = (data.get("uname") or "").strip()
    upass = (data.get("upass") or "").strip()
    if not uname or not upass:
        return JsonResponse({"error": "Missing credentials"}, status=400)
    try:
        bluconsole.blu_login(uname, upass)
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=401)
    request.session["blu_creds"] = {"uname": uname, "upass": upass}
    return JsonResponse({"ok": True})


@require_http_methods(["POST"])
def api_blu_logout(request):
    request.session.pop("blu_creds", None)
    return JsonResponse({"ok": True})


@require_http_methods(["GET"])
def api_blu_status(request):
    creds = _blu_creds(request)
    return JsonResponse({"authenticated": bool(creds)})


@require_http_methods(["GET"])
def api_blu_devices(request):
    creds = _blu_creds(request)
    if not creds:
        return JsonResponse({"error": "Not authenticated"}, status=401)
    try:
        xml = bluconsole.get_devices(creds["uname"], creds["upass"], children=False)
        devices = parse_devices(xml)
        return JsonResponse({"devices": devices})
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=502)


@require_http_methods(["GET"])
def api_blu_measurements(request):
    creds = _blu_creds(request)
    if not creds:
        return JsonResponse({"error": "Not authenticated"}, status=401)
    device_id = request.GET.get("id") or None
    from_time = request.GET.get("fromTime")
    to_time = request.GET.get("toTime")
    include_all = request.GET.get("includeAll") == "true"
    try:
        xml = bluconsole.get_measurements(
            creds["uname"],
            creds["upass"],
            device_id=device_id,
            from_time=int(from_time) if from_time else None,
            to_time=int(to_time) if to_time else None,
            include_all=include_all,
        )
        points = parse_measurements(xml, device_id=device_id)
        return JsonResponse({"points": points})
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=502)


@require_http_methods(["POST"])
def api_signup(request):
    data = _json_body(request)
    email = (data.get("email") or "").strip().lower()
    if not email:
        return JsonResponse({"error": "Email is required"}, status=400)
    first_name = (data.get("firstName") or "").strip()
    last_name = (data.get("lastName") or "").strip()
    profile, _ = Profile.objects.get_or_create(owner_key=email)
    profile.email = email
    profile.first_name = first_name
    profile.last_name = last_name
    profile.save()
    _set_owner_key(request, email)
    return JsonResponse({"ok": True})


@require_http_methods(["GET", "POST"])
def api_profile(request):
    owner_key = _owner_key(request)
    if request.method == "GET":
        profile = Profile.objects.filter(owner_key=owner_key).first()
        if not profile:
            return JsonResponse({"profile": None})
        return JsonResponse(
            {
                "profile": {
                    "firstName": profile.first_name,
                    "lastName": profile.last_name,
                    "email": profile.email,
                    "photoDataUrl": profile.photo_data_url,
                }
            }
        )
    data = _json_body(request)
    profile, _ = Profile.objects.get_or_create(owner_key=owner_key)
    profile.first_name = (data.get("firstName") or "").strip()
    profile.last_name = (data.get("lastName") or "").strip()
    profile.email = (data.get("email") or profile.email or "").strip()
    profile.photo_data_url = data.get("photoDataUrl") or ""
    profile.save()
    return JsonResponse({"ok": True})


@require_http_methods(["GET", "POST"])
def api_notes(request):
    owner_key = _owner_key(request)
    if request.method == "GET":
        notes = (
            Note.objects.filter(owner_key=owner_key)
            .order_by("-updated_at", "-created_at")
            .values("id", "title", "body", "created_at", "updated_at")
        )
        return JsonResponse({"notes": list(notes)})
    data = _json_body(request)
    title = (data.get("title") or "").strip()
    body = (data.get("body") or "").strip()
    if not title or not body:
        return JsonResponse({"error": "Title and body are required"}, status=400)
    note = Note.objects.create(owner_key=owner_key, title=title, body=body)
    return JsonResponse({"note": {"id": note.id, "title": note.title, "body": note.body}})


@require_http_methods(["PUT", "DELETE"])
def api_note_detail(request, note_id: int):
    owner_key = _owner_key(request)
    note = Note.objects.filter(owner_key=owner_key, id=note_id).first()
    if not note:
        return JsonResponse({"error": "Not found"}, status=404)
    if request.method == "DELETE":
        note.delete()
        return JsonResponse({"ok": True})
    data = _json_body(request)
    note.title = (data.get("title") or note.title).strip()
    note.body = (data.get("body") or note.body).strip()
    note.updated_at = timezone.now()
    note.save(update_fields=["title", "body", "updated_at"])
    return JsonResponse({"ok": True})


@require_http_methods(["GET", "POST"])
def api_uploads(request):
    owner_key = _owner_key(request)
    if request.method == "GET":
        uploads = (
            UploadDataset.objects.filter(owner_key=owner_key)
            .order_by("-created_at")
            .values("id", "name", "row_count", "created_at")
        )
        return JsonResponse({"uploads": list(uploads)})
    if request.FILES.get("file"):
        up = request.FILES["file"]
        if up.size > 5 * 1024 * 1024:
            return JsonResponse({"error": "File too large (max 5 MB)"}, status=400)
        name = up.name
        lower = name.lower()
        if lower.endswith(".xls") and not lower.endswith(".xlsx"):
            return JsonResponse({"error": "Please save as .xlsx for now."}, status=400)
        if not lower.endswith(".xlsx"):
            return JsonResponse({"error": "Unsupported file type."}, status=400)
        if openpyxl is None:
            return JsonResponse({"error": "Excel parser not available."}, status=500)
        try:
            dataset = _parse_xlsx_dataset(up.read())
        except Exception as exc:  # noqa: BLE001
            return JsonResponse({"error": str(exc)}, status=400)
        try:
            upload = UploadDataset.objects.create(
                owner_key=owner_key,
                name=name,
                headers=dataset["headers"],
                rows=dataset["rows"],
                row_count=len(dataset["rows"]),
            )
        except Exception as exc:  # noqa: BLE001
            return JsonResponse({"error": f"Unable to save dataset: {exc}"}, status=400)
        return JsonResponse(
            {
                "upload": {
                    "id": upload.id,
                    "name": upload.name,
                    "headers": upload.headers,
                    "rows": upload.rows,
                    "row_count": upload.row_count,
                    "created_at": upload.created_at,
                }
            }
        )
    data = _json_body(request)
    name = (data.get("name") or "").strip()
    headers = data.get("headers") or []
    rows = data.get("rows") or []
    if not name:
        return JsonResponse({"error": "Name is required"}, status=400)
    upload = UploadDataset.objects.create(
        owner_key=owner_key,
        name=name,
        headers=headers,
        rows=rows,
        row_count=len(rows),
    )
    return JsonResponse({"upload": {"id": upload.id}})


@require_http_methods(["DELETE"])
def api_uploads_clear(request):
    owner_key = _owner_key(request)
    UploadDataset.objects.filter(owner_key=owner_key).delete()
    return JsonResponse({"ok": True})


@require_http_methods(["GET", "DELETE"])
def api_upload_detail(request, upload_id: int):
    owner_key = _owner_key(request)
    upload = UploadDataset.objects.filter(owner_key=owner_key, id=upload_id).first()
    if not upload:
        return JsonResponse({"error": "Not found"}, status=404)
    if request.method == "DELETE":
        upload.delete()
        return JsonResponse({"ok": True})
    return JsonResponse(
        {
            "upload": {
                "id": upload.id,
                "name": upload.name,
                "headers": upload.headers,
                "rows": upload.rows,
                "row_count": upload.row_count,
                "created_at": upload.created_at,
            }
        }
    )


@require_http_methods(["POST"])
def api_ai_chat(request):
    if not settings.OPENAI_API_KEY:
        return JsonResponse({"error": "OPENAI_API_KEY not configured"}, status=400)
    data = _json_body(request)
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return JsonResponse({"error": "Prompt is required"}, status=400)
    session_id = data.get("session_id")
    attachment = data.get("attachment") or None
    session = _get_or_create_chat_session(request, session_id, prompt)
    user_msg = ChatMessage.objects.create(session=session, role="user", content=prompt)
    if attachment:
        ChatAttachment.objects.create(
            message=user_msg,
            name=str(attachment.get("name") or "attachment"),
            mime=str(attachment.get("mime") or ""),
            summary=attachment.get("summary") or {},
        )
    try:
        context = _build_ai_context(request, prompt, session)
        answer = _openai_chat(prompt, context, session)
        if not answer:
            answer = _fallback_attachment_answer(session) or "I can help. What specific insight do you need?"
        if _looks_like_blind_reply(answer):
            answer = _fallback_attachment_answer(session) or answer
        ChatMessage.objects.create(session=session, role="assistant", content=answer)
        session.updated_at = timezone.now()
        session.save(update_fields=["updated_at"])
        return JsonResponse({"answer": answer, "session_id": session.id})
    except Exception as exc:  # noqa: BLE001
        fallback = _fallback_attachment_answer(session)
        if fallback:
            ChatMessage.objects.create(session=session, role="assistant", content=fallback)
            session.updated_at = timezone.now()
            session.save(update_fields=["updated_at"])
            return JsonResponse({"answer": fallback, "session_id": session.id})
        return JsonResponse({"error": str(exc), "session_id": session.id}, status=502)


@require_http_methods(["GET"])
def api_ai_chat_status(request):
    return JsonResponse(
        {
            "connected": bool(settings.OPENAI_API_KEY),
            "provider": "openai",
        }
    )


@require_http_methods(["GET"])
def api_ai_chat_sessions(request):
    owner_key = _owner_key(request)
    sessions = (
        ChatSession.objects.filter(owner_key=owner_key)
        .order_by("-updated_at")
        .values("id", "title", "created_at", "updated_at")
    )
    return JsonResponse({"sessions": list(sessions)})


@require_http_methods(["GET"])
def api_ai_chat_session_detail(request, session_id: int):
    owner_key = _owner_key(request)
    session = ChatSession.objects.filter(owner_key=owner_key, id=session_id).first()
    if not session:
        return JsonResponse({"error": "Not found"}, status=404)
    messages = (
        ChatMessage.objects.filter(session=session)
        .order_by("created_at")
        .values("id", "role", "content", "created_at")
    )
    return JsonResponse({"session": {"id": session.id, "title": session.title}, "messages": list(messages)})


@require_http_methods(["DELETE"])
def api_ai_chat_sessions_clear(request):
    owner_key = _owner_key(request)
    ChatSession.objects.filter(owner_key=owner_key).delete()
    return JsonResponse({"ok": True})


@require_http_methods(["POST"])
def api_ai_chat_attachment(request):
    up = request.FILES.get("file")
    if not up:
        return JsonResponse({"error": "Missing file"}, status=400)
    if up.size > 5 * 1024 * 1024:
        return JsonResponse({"error": "File too large (max 5 MB)"}, status=400)

    name = up.name
    mime = up.content_type or ""
    if mime.startswith("image/"):
        return JsonResponse(
            {"attachment": {"name": name, "mime": mime, "summary": {"type": "image", "name": name, "size": up.size}}}
        )

    lower = name.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return JsonResponse({"error": "Please save as .xlsx for now."}, status=400)
    if not lower.endswith(".xlsx"):
        return JsonResponse({"error": "Unsupported file type."}, status=400)
    if openpyxl is None:
        return JsonResponse({"error": "Excel parser not available."}, status=500)

    try:
        summary = _summarize_xlsx(up.read())
        return JsonResponse({"attachment": {"name": name, "mime": mime, "summary": summary}})
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=400)


def _extract_logger_id(prompt: str) -> str | None:
    matches = re.findall(r"\b\d{3,}\b", prompt)
    return matches[0] if matches else None


def _wants_logger_status(prompt: str) -> bool:
    text = prompt.lower()
    keywords = ["status", "online", "offline", "logger", "loggers", "battery", "alert", "alerts"]
    return any(k in text for k in keywords)


def _get_logger_status_snapshot(creds: dict) -> str:
    try:
        xml = bluconsole.get_devices(creds["uname"], creds["upass"], children=False)
        devices = parse_devices(xml)
    except Exception:
        return "Logger status: unable to fetch device list right now."

    if not devices:
        return "Logger status: no devices found."

    now = int(datetime.now(tz=dt_timezone.utc).timestamp())
    from_time = now - 48 * 3600
    sample = devices[:25]
    results = {}

    def fetch_latest(dev):
        try:
            xml_m = bluconsole.get_measurements(
                creds["uname"],
                creds["upass"],
                device_id=str(dev.get("id")),
                from_time=from_time,
                to_time=now,
            )
            points = parse_measurements(xml_m, device_id=str(dev.get("id")))
            latest = None
            for p in points:
                if not latest or (p.get("utc") or 0) > (latest.get("utc") or 0):
                    latest = p
            return dev, latest
        except Exception:
            return dev, None

    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = [pool.submit(fetch_latest, d) for d in sample if d.get("id")]
        for f in as_completed(futures):
            dev, latest = f.result()
            results[str(dev.get("id"))] = latest

    online = 0
    offline = 0
    low_batt = 0
    details = []
    for d in sample:
        did = str(d.get("id"))
        latest = results.get(did)
        last_utc = latest.get("utc") if latest else None
        minutes = (now - last_utc) / 60 if last_utc else None
        is_online = minutes is not None and minutes < 30
        if is_online:
            online += 1
        else:
            offline += 1
        if d.get("battery") is not None and d.get("battery") < 20:
            low_batt += 1
        details.append(
            f"{did} ({d.get('label') or 'no label'}): "
            f"{'online' if is_online else 'offline'}, "
            f"last={_format_dt_from_utc(last_utc)}, "
            f"temp={latest.get('t') if latest else 'n/a'}"
        )

    summary = (
        f"Logger status snapshot (sample {len(sample)} of {len(devices)}): "
        f"online {online}, offline {offline}, low battery {low_batt}."
    )
    return summary + " Recent loggers: " + "; ".join(details[:8]) + "."


def _format_dt_from_utc(utc: int | None) -> str:
    if not utc:
        return "unknown"
    return datetime.fromtimestamp(utc, tz=dt_timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _build_ai_context(request, prompt: str, session: ChatSession | None) -> str:
    owner_key = _owner_key(request)
    profile = Profile.objects.filter(owner_key=owner_key).first()
    notes = (
        Note.objects.filter(owner_key=owner_key)
        .order_by("-updated_at", "-created_at")
        .values("title", "body", "updated_at", "created_at")[:5]
    )
    uploads = (
        UploadDataset.objects.filter(owner_key=owner_key)
        .order_by("-created_at")
        .values("name", "row_count", "created_at")[:5]
    )

    context_lines = [
        "Dashboard pages: Home, Sensor Feed (devices + uploads), Visualizations (charts), AI, FAQ, Profile, Contact.",
    ]
    if profile:
        context_lines.append(
            f"Profile: {profile.first_name} {profile.last_name} ({profile.email or 'no email'})"
        )
    if notes:
        context_lines.append("Recent notes:")
        for n in notes:
            when = n["updated_at"] or n["created_at"]
            context_lines.append(f"- {n['title']}: {n['body']} (at {when})")
    if uploads:
        context_lines.append("Recent uploads:")
        for u in uploads:
            context_lines.append(f"- {u['name']} ({u['row_count']} rows, {u['created_at']})")

    creds = _blu_creds(request)
    if creds:
        logger_id = _extract_logger_id(prompt)
        if logger_id:
            try:
                xml = bluconsole.get_devices(creds["uname"], creds["upass"], children=False)
                devices = parse_devices(xml)
                device = next((d for d in devices if str(d.get("id")) == str(logger_id)), None)
                if device:
                    context_lines.append(
                        "Logger info: "
                        f"id={device.get('id')}, type={device.get('type')}, "
                        f"label={device.get('label')}, min={device.get('min_temp')}, "
                        f"max={device.get('max_temp')}, vrn={device.get('vrn')}"
                    )
                now = int(datetime.now(tz=dt_timezone.utc).timestamp())
                from_time = now - 48 * 3600
                xml_m = bluconsole.get_measurements(
                    creds["uname"],
                    creds["upass"],
                    device_id=str(logger_id),
                    from_time=from_time,
                    to_time=now,
                )
                points = parse_measurements(xml_m, device_id=str(logger_id))
                latest = None
                for p in points:
                    if not latest or (p.get("utc") or 0) > (latest.get("utc") or 0):
                        latest = p
                if latest:
                    context_lines.append(
                        "Latest measurement: "
                        f"temp={latest.get('t')}, hum={latest.get('h')}, "
                        f"utc={_format_dt_from_utc(latest.get('utc'))}"
                    )
                else:
                    context_lines.append("Latest measurement: none found in last 48h.")
            except Exception:
                context_lines.append("BluConsole: unable to fetch logger data right now.")
        else:
            context_lines.append(
                "BluConsole: user connected. You can ask about trends, alerts, or provide a logger id for live data."
            )
            if _wants_logger_status(prompt):
                context_lines.append(_get_logger_status_snapshot(creds))
    else:
        context_lines.append("BluConsole: user not connected.")

    if session:
        last_msgs = (
            ChatMessage.objects.filter(session=session)
            .order_by("-created_at")
            .values("role", "content")[:6]
        )
        if last_msgs:
            context_lines.append("Recent chat:")
            for m in reversed(list(last_msgs)):
                context_lines.append(f"- {m['role']}: {m['content']}")
        last_attach = (
            ChatAttachment.objects.filter(message__session=session)
            .order_by("-created_at")
            .values("name", "mime", "summary")
            .first()
        )
        if last_attach:
            context_lines.append(f"Latest attachment: {last_attach['name']} ({last_attach['mime']})")
            context_lines.append(f"Attachment summary: {last_attach['summary']}")
            summary = last_attach.get("summary") or {}
            if summary.get("type") == "excel":
                context_lines.append(f"Attachment columns: {summary.get('headers')}")
                context_lines.append(
                    "Attachment insights: "
                    f"rows={summary.get('rows')}, time={summary.get('timeStart')} to {summary.get('timeEnd')}, "
                    f"temp min={summary.get('tempMin')}, max={summary.get('tempMax')}, avg={summary.get('tempAvg')}"
                )
            elif summary:
                context_lines.append(f"Attachment summary fields: {list(summary.keys())}")

    return "\n".join(context_lines) or "No local data found."


def _openai_chat(prompt: str, context: str, session: ChatSession | None) -> str:
    system_msg = (
        "You are an expert poultry cold-chain analyst and dashboard assistant. "
        "Use the provided context (including attachment summaries) to deliver clear, structured analysis. "
        "Never claim you cannot access an attached file if a summary is present. "
        "If data is missing, ask one clarifying question but still give general guidance. "
        "Write in short sections with headings and bullet points, like:\n"
        "1) What this file contains\n"
        "2) Time vs Temperature (what happened)\n"
        "3) Temperature abuse (yes/no, why)\n"
        "4) Shelf-life impact (plain language)\n"
        "5) Practical interpretation\n"
        "6) One-line summary\n"
        "When the user greets you or asks a general question, respond warmly and ask what they want to analyze."
    )
    messages = [
        {"role": "system", "content": system_msg},
        {"role": "system", "content": f"Context:\n{context}"},
        {
            "role": "system",
            "content": (
                "Format your response as numbered sections with short headings and bullet points. "
                "Use markdown. Keep each section short and clear."
            ),
        },
    ]
    if session:
        history = (
            ChatMessage.objects.filter(session=session)
            .order_by("-created_at")
            .values("role", "content")[:8]
        )
        for m in reversed(list(history)):
            if m["role"] in ("user", "assistant"):
                messages.append({"role": m["role"], "content": m["content"]})
    messages.append({"role": "user", "content": prompt})

    payload = {
        "model": settings.OPENAI_MODEL,
        "messages": messages,
        "temperature": 0.4,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urlrequest.Request(
        "https://api.openai.com/v1/chat/completions",
        data=data,
        headers={
            "Authorization": f"Bearer {settings.OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urlrequest.urlopen(req, timeout=25) as resp:
        body = json.loads(resp.read().decode("utf-8"))
    choice = body.get("choices", [{}])[0]
    message = choice.get("message", {})
    content = (message.get("content") or "").strip() or "No response from model."
    if _needs_structure(content):
        return _format_structured_answer(session, prompt)
    return _normalize_structured_text(content)


def _get_or_create_chat_session(request, session_id: Any, prompt: str) -> ChatSession:
    owner_key = _owner_key(request)
    session = None
    if session_id:
        session = ChatSession.objects.filter(owner_key=owner_key, id=session_id).first()
    if session:
        return session
    title = prompt.strip().splitlines()[0][:60] or "New chat"
    return ChatSession.objects.create(owner_key=owner_key, title=title)


def _fallback_attachment_answer(session: ChatSession | None) -> str | None:
    if not session:
        return None
    last_attach = (
        ChatAttachment.objects.filter(message__session=session)
        .order_by("-created_at")
        .values("summary", "name")
        .first()
    )
    if not last_attach:
        return None
    summary = last_attach.get("summary") or {}
    if summary.get("type") != "excel":
        return "I can see an attachment, but it is not an Excel file. Please upload a logger export (.xls/.xlsx)."
    time_start = summary.get("timeStart") or "unknown"
    time_end = summary.get("timeEnd") or "unknown"
    t_min = summary.get("tempMin")
    t_max = summary.get("tempMax")
    t_avg = summary.get("tempAvg")
    t_first = summary.get("tempAtStart")
    t_last = summary.get("tempAtEnd")
    t_samples = summary.get("tempSamples")
    return (
        "From the uploaded logger file: "
        f"time range {time_start} to {time_end}; "
        f"temperature min={t_min}, max={t_max}, avg={t_avg}, "
        f"start={t_first}, end={t_last}, samples={t_samples}. "
        "I can explain trends, anomalies, and shelf-life insights from this."
    )


def _looks_like_blind_reply(answer: str) -> bool:
    text = answer.lower()
    triggers = [
        "can't access",
        "cannot access",
        "do not have access",
        "provide a logger id",
        "please provide the logger id",
        "share the logger id",
    ]
    return any(t in text for t in triggers)


def _needs_structure(answer: str) -> bool:
    text = answer.strip().lower()
    if not text:
        return True
    markers = ["1)", "1.", "2)", "2.", "one-line summary", "one line summary"]
    return not any(m in text for m in markers)


def _normalize_structured_text(text: str) -> str:
    if "\n" in text:
        return text
    # Insert line breaks before numbered sections and bullets for readability.
    normalized = re.sub(r"\s(\d\))", r"\n\n\1", text)
    normalized = re.sub(r"\s-\s", r"\n- ", normalized)
    return normalized.strip()


def _format_structured_answer(session: ChatSession | None, prompt: str) -> str:
    fallback = _fallback_attachment_answer(session)
    summary = None
    if session:
        last_attach = (
            ChatAttachment.objects.filter(message__session=session)
            .order_by("-created_at")
            .values("summary")
            .first()
        )
        summary = last_attach.get("summary") if last_attach else None

    if not summary or summary.get("type") != "excel":
        return (
            "1) What this file contains\n"
            "- I do not have a parsed Excel summary yet.\n"
            "\n"
            "2) Time vs Temperature (what happened)\n"
            "- Please attach a .xlsx logger export so I can compute min/max/avg and time range.\n"
            "\n"
            "3) Temperature abuse (yes/no, why)\n"
            "- Cannot assess without data.\n"
            "\n"
            "4) Shelf-life impact\n"
            "- In general, higher temperatures accelerate degradation (Arrhenius effect).\n"
            "\n"
            "5) Practical interpretation\n"
            "- Re-attach the file or mention a logger ID for live data.\n"
            "\n"
            "6) One-line summary\n"
            "- Upload the file and I will summarize it in detail."
        )

    time_start = summary.get("timeStart") or "unknown"
    time_end = summary.get("timeEnd") or "unknown"
    t_min = summary.get("tempMin")
    t_max = summary.get("tempMax")
    t_avg = summary.get("tempAvg")
    t_first = summary.get("tempAtStart")
    t_last = summary.get("tempAtEnd")
    rows = summary.get("rows")

    return (
        "1) What this file contains\n"
        f"- {rows} readings covering {time_start} to {time_end}\n"
        "- Temperature readings from a logger export\n"
        "\n"
        "2) Time vs Temperature (what happened)\n"
        f"- Min: {t_min} C, Max: {t_max} C, Avg: {t_avg} C\n"
        f"- Start: {t_first} C, End: {t_last} C\n"
        "\n"
        "3) Temperature abuse (yes/no, why)\n"
        "- If temperatures repeatedly rise above your target cold-chain range, that is thermal stress.\n"
        "\n"
        "4) Shelf-life impact (plain language)\n"
        "- Higher temperatures consume shelf life faster; small increases add up over time.\n"
        "\n"
        "5) Practical interpretation\n"
        "- If this pattern repeats, expect shorter shelf life and higher quality risk.\n"
        "\n"
        "6) One-line summary\n"
        f"- Temperatures stayed between {t_min} and {t_max} C from {time_start} to {time_end}, "
        "which can gradually reduce shelf life if sustained."
    )


def _summarize_xlsx(raw: bytes) -> dict:
    wb = _load_workbook_safe(raw)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"type": "excel", "rows": 0}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = rows[1:]

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, (int, float)):
            if 20_000 < val < 60_000:
                return datetime(1899, 12, 30, tzinfo=dt_timezone.utc) + timedelta(days=float(val))
            if val > 1_000_000_000:
                return datetime.fromtimestamp(val, tz=dt_timezone.utc)
        try:
            d = datetime.fromisoformat(str(val))
            return d
        except Exception:
            return None

    def to_num(val):
        try:
            return float(val)
        except Exception:
            return None

    # detect time/temp columns
    time_col = ""
    temp_col = ""
    best_time = -1
    best_temp = -1
    sample = body[:200]
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        score = 0
        if re.search(r"(time|date|timestamp)", h_lower):
            score += 2
        for r in sample:
            if idx < len(r) and parse_date(r[idx]):
                score += 1
        if score > best_time:
            best_time = score
            time_col = h
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        score = 0
        if re.search(r"(temp|temperature|degc|celsius)", h_lower):
            score += 2
        ok = 0
        tot = 0
        for r in sample:
            if idx < len(r):
                if to_num(r[idx]) is not None:
                    ok += 1
                tot += 1
        if tot:
            score += (ok / tot) * 5
        if score > best_temp:
            best_temp = score
            temp_col = h

    time_idx = headers.index(time_col) if time_col in headers else None
    temp_idx = headers.index(temp_col) if temp_col in headers else None

    t_min = None
    t_max = None
    t_sum = 0.0
    t_count = 0
    t_start = None
    t_end = None
    t_first = None
    t_last = None

    for r in body:
        tval = None
        dval = None
        if temp_idx is not None and temp_idx < len(r):
            tval = to_num(r[temp_idx])
        if time_idx is not None and time_idx < len(r):
            dval = parse_date(r[time_idx])
        if tval is not None:
            t_min = tval if t_min is None else min(t_min, tval)
            t_max = tval if t_max is None else max(t_max, tval)
            t_sum += tval
            t_count += 1
        if dval:
            if t_start is None or dval < t_start:
                t_start = dval
                t_first = tval
            if t_end is None or dval > t_end:
                t_end = dval
                t_last = tval

    return {
        "type": "excel",
        "rows": len(body),
        "headers": headers,
        "timeCol": time_col,
        "tempCol": temp_col,
        "tempMin": t_min,
        "tempMax": t_max,
        "tempAvg": (t_sum / t_count) if t_count else None,
        "timeStart": t_start.isoformat() if t_start else None,
        "timeEnd": t_end.isoformat() if t_end else None,
        "tempAtStart": t_first,
        "tempAtEnd": t_last,
        "tempSamples": t_count,
    }


def _serialize_cell(val: Any):
    if val is None:
        return None
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    if isinstance(val, timedelta):
        return str(val)
    if isinstance(val, (bytes, bytearray)):
        return val.decode("utf-8", errors="replace")
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)


def _parse_xlsx_dataset(raw: bytes) -> dict:
    wb = _load_workbook_safe(raw)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"headers": [], "rows": []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = []
    for r in rows[1:]:
        obj = {}
        for i, h in enumerate(headers):
            val = r[i] if i < len(r) else None
            obj[h] = _serialize_cell(val)
        body.append(obj)
    return {"headers": headers, "rows": body}


def _load_workbook_safe(raw: bytes):
    try:
        return openpyxl.load_workbook(
            filename=BytesIO(raw),
            data_only=True,
            read_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Unable to read workbook. Please re-save the file as .xlsx (new copy) and try again."
        ) from exc
